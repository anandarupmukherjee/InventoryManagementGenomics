import io
import csv
import zipfile
import datetime
from django.apps import apps
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import make_aware
from openpyxl import Workbook
from django.utils.timezone import now

from services.data_storage.models import Product, Withdrawal, PurchaseOrder  # Adjust as needed



MODEL_MAP = {
    'Withdrawal': 'data_storage.Withdrawal',
    'Product': 'data_storage.Product',
    'PurchaseOrder': 'data_storage.PurchaseOrder',
}


FILTER_FIELDS = {
    'Withdrawal': 'timestamp',
    'PurchaseOrder': 'order_date',
    # Product has no date field to filter
}

def download_report(request):
    selected_model = request.GET.get('model', 'Withdrawal')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    download_type = request.GET.get('download')

    preview_model_class = apps.get_model(*MODEL_MAP[selected_model].split('.'))
    preview_fields = [f.name for f in preview_model_class._meta.fields]

    preview_queryset = preview_model_class.objects.all()

    # Filter preview table if model supports date filtering
    if selected_model in FILTER_FIELDS:
        date_field = FILTER_FIELDS[selected_model]
        if start_date:
            preview_queryset = preview_queryset.filter(
                **{f"{date_field}__gte": make_aware(datetime.datetime.strptime(start_date, "%Y-%m-%d"))}
            )
        if end_date:
            preview_queryset = preview_queryset.filter(
                **{f"{date_field}__lte": make_aware(datetime.datetime.strptime(end_date, "%Y-%m-%d"))}
            )

    # ✅ THEN slice it
    preview_queryset = preview_queryset.order_by('-id')


    # Handle Excel or CSV download
    if download_type in ['excel', 'csv']:
        model_class = preview_model_class
        fields = preview_fields
        qs = preview_queryset

        # Build filename
        user_part = request.user.username if request.user.is_authenticated else "anonymous"
        date_part = now().strftime("%Y%m%d")

        if download_type == 'excel':
            wb = Workbook()
            ws = wb.create_sheet(title=selected_model)
            ws.append(fields)
            for obj in qs:
                ws.append([str(getattr(obj, field, '')) for field in fields])

            # Remove the default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            response = HttpResponse(out, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=IMS_{user_part}_{date_part}.xlsx'
            return response

        elif download_type == 'csv':
            csv_buffer = io.StringIO()
            writer = csv.writer(csv_buffer)
            writer.writerow(fields)
            for obj in qs:
                writer.writerow([str(getattr(obj, field, '')) for field in fields])

            csv_buffer.seek(0)
            response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename=IMS_{user_part}_{date_part}.csv'
            return response

    return render(request, 'analytics/download_report.html', {
        'models': MODEL_MAP.keys(),
        'selected_model': selected_model,
        'fields': preview_fields,
        'data': preview_queryset,
        'start_date': start_date,
        'end_date': end_date,
    })
